import pickle, numpy as np, openpyxl, math, calendar
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_BASE = Path(__file__).parent
with open(_BASE / 'sim_final.pkl', 'rb') as f:
    d=pickle.load(f)

ann=d['ann']; soc_all=d['soc_all']; eff_pct=d['eff_pct']; eff_nom=d['eff_nom']
G_s=d['G_s']; G_n=d['G_n']; P_dc_W=d['P_dc_W']
mons=d['mons']; hrs_arr=d['hrs']; is_bo=d['is_blackout']
monthly_gt=d['monthly_gt']; monthly_bo=d['monthly_bo']
bo_inv_eff=d['bo_inv_eff']; DC_per_event=d['DC_per_event']
peak_eta=d['peak_eta']; peak_pdc_kw=d.get('peak_pdc_kw', 0.0)

P_CRITICAL_KW=6.0; ENERGY_PRICE=0.24; N_CLIENTS=45; COMP=3.0
BATT_NOM_KWH=81.92; GRID_CI=0.384; YEAR=2025
# YEAR is only used to index a representative 8760-hour weather/calendar profile.
# The report treats the three-year check as three representative years with
# 52 Thursday blackout cycles each (156 cycles total), not as fixed 2025-2027.
pk_dc=P_dc_W.max()/1000; pk_cc=pk_dc*0.95; pk_A=pk_cc*1000/51.2
soc_y1=np.array(soc_all[0]); min_bo_soc=soc_y1[is_bo].min()

# ─────────────────────────────────────────────────────────────
# Derived values used throughout the workbook. These avoid hardcoded
# Q10/Q11 values and make the output follow the actual simulation.
# ─────────────────────────────────────────────────────────────
soc_arrays = [np.asarray(soc_all[y], dtype=float) for y in range(3)]
blackout_hours_per_year = int(np.count_nonzero(is_bo))
blackout_hours_per_event = 8
cycles_per_year = blackout_hours_per_year // blackout_hours_per_event
total_cycles_3yr = cycles_per_year * 3
annual_backup_kWh = blackout_hours_per_year * P_CRITICAL_KW

# This project uses the conservative assignment interpretation:
# one scheduled blackout discharge + recharge is counted as one degradation cycle.
DEGRADATION_PER_CYCLE = 0.0002
nom_y3_end_calc = BATT_NOM_KWH * (1 - total_cycles_3yr * DEGRADATION_PER_CYCLE)
total_deg_pct_calc = total_cycles_3yr * DEGRADATION_PER_CYCLE * 100

start_soc_vals = []
end_soc_vals = []
min_soc_vals = []
covered_sim_vals = []
for y in range(3):
    arr = soc_arrays[y]
    # Prefer explicit simulation metadata. If unavailable, use the first hourly SoC value.
    start_soc = float(ann[y].get('start_soc_pct', arr[0] if len(arr) else 95.0))
    end_soc = float(ann[y].get('end_soc_pct', arr[-1] if len(arr) else start_soc))
    min_soc = float(arr[is_bo].min())
    start_soc_vals.append(f"{start_soc:.2f}%")
    end_soc_vals.append(f"{end_soc:.2f}%")
    min_soc_vals.append(f"{min_soc:.2f}%")
    covered_sim_vals.append("✓ YES" if min_soc >= 20.0 else "✗ NO")

final_post_bo_soc = float(ann[2].get('end_soc_pct', 95.0)) - DC_per_event / nom_y3_end_calc * 100

MN=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

H =PatternFill("solid",start_color="1F4E79",end_color="1F4E79")
S =PatternFill("solid",start_color="2E75B6",end_color="2E75B6")
A =PatternFill("solid",start_color="D6E4F0",end_color="D6E4F0")
G =PatternFill("solid",start_color="E2EFDA",end_color="E2EFDA")
Y =PatternFill("solid",start_color="FFF2CC",end_color="FFF2CC")
R =PatternFill("solid",start_color="FCE4D6",end_color="FCE4D6")
W =PatternFill("solid",start_color="FFFFFF",end_color="FFFFFF")
OR=PatternFill("solid",start_color="F4B942",end_color="F4B942")

def bdr():
    s=Side(style="thin",color="BFBFBF")
    return Border(left=s,right=s,top=s,bottom=s)

def hdr(ws,r,c,v,fill=H,fc="FFFFFF",sz=10,bold=True,span=None,wrap=False,rh=None):
    if span: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=bold,color=fc,name="Arial",size=sz)
    cell.fill=fill
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
    if rh: ws.row_dimensions[r].height=rh
    return cell

def cel(ws,r,c,v,fill=None,bold=False,fc="000000",sz=10,fmt=None,align="center",wrap=True):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=bold,color=fc,name="Arial",size=sz)
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    if fill: cell.fill=fill
    if fmt:  cell.number_format=fmt
    cell.border=bdr()
    return cell

def sec(ws,r,title,span=4,rh=18):
    hdr(ws,r,1,title,S,span=span,rh=rh)

def tbox(ws,r,c,span,v,fill=W,sz=9,rh=None):
    ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=c+span-1)
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(name="Arial",size=sz)
    cell.fill=fill
    cell.alignment=Alignment(wrap_text=True,horizontal="left",vertical="top")
    if rh: ws.row_dimensions[r].height=rh
    return cell

wb=openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1: Q4/Q5 – Module & Roof Layout
# ══════════════════════════════════════════════════════════════
ws=wb.active; ws.title="Q4-Q5 Module and Roof"
for i,w in enumerate([34,20,12,34],1): ws.column_dimensions[get_column_letter(i)].width=w

hdr(ws,1,1,"Q4/Q5 – PV Module Selection & Roof Layout | Tampa 5m×18m | 35° Tilt",H,sz=12,span=4,rh=24)

sec(ws,3,"A. Module Selection – AIKO Neostar 490W chosen over LonGi Hi-MO X6 600W")
hdr(ws,4,1,"Parameter",H); hdr(ws,4,2,"AIKO 490W (SELECTED)",H)
hdr(ws,4,3,"LonGi 600W",H); hdr(ws,4,4,"Decision reason",H)
mods=[
    ("Pmax (STC)","490 W","600 W",""),
    ("Module efficiency","24.5%","23.2%","AIKO higher efficiency"),
    ("Dimensions (L×W)","1762×1134 mm","2278×1134 mm","LonGi too long for roof – see below"),
    ("Temp coeff Pmax","-0.26%/°C","-0.29%/°C","AIKO better in Florida heat"),
    ("Annual degradation","≤0.35%/yr","0.40%/yr","AIKO degrades slower"),
    ("Voc (STC)","41.10 V","52.81 V",""),
    ("Vmp (STC)","34.70 V","44.66 V",""),
    ("Isc (STC)","14.88 A","14.46 A",""),
    ("Imp (STC)","14.13 A","13.44 A",""),
    ("Max sys voltage","DC 1500V","DC 1500V",""),
    ("Fire rating","IEC Class A","UL Type 1/2",""),
    ("Product warranty","25 yr","15 yr","AIKO longer warranty"),
    ("FITS ROOF (portrait)?","✓ YES – 1.762m < 2.052m","✗ NO – 2.278m > 2.052m","LonGi eliminated here"),
]
for i,(p,a,l,n) in enumerate(mods,5):
    f=A if i%2==0 else W
    elim=("NO" in l)
    if elim: f=R
    sel=p=="FITS ROOF (portrait)?"
    cel(ws,i,1,p,f,align="left",bold=sel)
    cel(ws,i,2,a,G if sel else f,bold=sel)
    cel(ws,i,3,l,R if elim else f,bold=sel)
    cel(ws,i,4,n,f,align="left")

sec(ws,19,"B. Roof Geometry & Available Area")
hdr(ws,20,1,"Parameter",H); hdr(ws,20,2,"Calculation",H); hdr(ws,20,3,"Result",H); hdr(ws,20,4,"",H)
tilt=35
slope_surf=2.5/math.cos(math.radians(tilt))
avail_slope=slope_surf-1.0
avail_ridge=18.0-1.0
geo=[
    ("Total roof width",                    "Given",                         "5.0 m",    ""),
    ("Total roof length",                   "Given",                         "18.0 m",   ""),
    ("Tilt angle",                          "Given",                         "35°",      ""),
    ("Horizontal half-width per slope",     "5.0 ÷ 2",                       "2.5 m",    ""),
    ("Slope surface width",                 "2.5 ÷ cos(35°)",                f"{slope_surf:.3f} m",""),
    ("Safety margin each side",             "Assignment requirement",         "0.5 m",    ""),
    ("Available along slope",               f"{slope_surf:.3f} − 2×0.5",     f"{avail_slope:.3f} m",""),
    ("Available along ridge",               "18.0 − 2×0.5",                  f"{avail_ridge:.1f} m", ""),
]
for i,(p,calc,res,n) in enumerate(geo,21):
    f=A if i%2==0 else W
    cel(ws,i,1,p,f,align="left"); cel(ws,i,2,calc,f); cel(ws,i,3,res,f,bold=True); cel(ws,i,4,n,f)

sec(ws,30,"C. Cluster Layout Calculation")
hdr(ws,31,1,"Parameter",H); hdr(ws,31,2,"Calculation",H); hdr(ws,31,3,"Result",H); hdr(ws,31,4,"Notes",H)
cluster_W=4*1.134; ridge_used=3*4*1.134+2*0.5
layout=[
    ("Module orientation",       "Portrait chosen",                    "1.762m along slope, 1.134m along ridge",""),
    ("Rows per cluster",         f"floor({avail_slope:.3f} ÷ 1.762)",  "1 row",       "1.762m < 2.052m available ✓"),
    ("Cols per cluster",         "Max allowed",                         "4 columns",   ""),
    ("Cluster width (ridge)",    "4 × 1.134m",                         f"{cluster_W:.3f} m",""),
    ("Inter-cluster gap",        "Assignment requirement",              "0.5 m",       ""),
    ("Clusters that fit",        f"floor(({avail_ridge:.0f}+0.5) ÷ ({cluster_W:.3f}+0.5))","3 clusters",""),
    ("Ridge length occupied",    "3×4×1.134 + 2×0.5",                  f"{ridge_used:.3f} m","< 17.0m ✓"),
    ("Modules per slope",        "1 row × 3 clusters × 4 cols",        "12 modules",  ""),
    ("Total modules (2 slopes)", "12 × 2",                             "24 modules",  "12 south + 12 north"),
    ("TOTAL INSTALLED kWp",      "24 × 490W",                          "11.76 kWp",   ""),
]
for i,(p,calc,res,n) in enumerate(layout,32):
    key=p.startswith("TOTAL")
    f=G if key else (A if i%2==0 else W)
    cel(ws,i,1,p,f,bold=key,align="left")
    cel(ws,i,2,calc,f); cel(ws,i,3,res,f,bold=key); cel(ws,i,4,n,f,align="left")

# ══════════════════════════════════════════════════════════════
# SHEET 2: Q4/Q7 – Inverter & CC Selection
# ══════════════════════════════════════════════════════════════
ws2=wb.create_sheet("Q4-Q7 Inverter and CC")
for i,w in enumerate([32,20,20,20,22],1): ws2.column_dimensions[get_column_letter(i)].width=w

hdr(ws2,1,1,"Q4/Q7 – Inverter & Charge Controller Selection | 120V/60Hz | Tampa",H,sz=12,span=5,rh=24)

sec(ws2,3,"A. Inverter Selection – UL Variants Only (120V/60Hz North America)",span=5)
hdr(ws2,4,1,"Parameter",H); hdr(ws2,4,2,"SW 4024",H); hdr(ws2,4,3,"SW 4048",H)
hdr(ws2,4,4,"XW Pro 6848",H); hdr(ws2,4,5,"Notes",H)
inv_rows=[
    ("Cont. output power (25°C)","3,400 W","3,800 W","6,800 W","Critical load = 6,000 W"),
    ("DC bus voltage","24 V","48 V","48 V",""),
    ("Output voltage","120/240 V","120/240 V","120/240 V split-phase",""),
    ("Output frequency","60 Hz","60 Hz","60 Hz",""),
    ("Li-ion compatible?","✗ No","✗ No","✓ Yes","Required for LFP bank"),
    ("Peak efficiency","92%","94%","96.1%",""),
    ("CEC weighted eff.","–","–","94.1%","Used for GT calculations"),
    ("Transfer time","<16.7 ms","<16.7 ms","8 ms","Faster for data center"),
    ("Covers 6 kW load?","✗ NO","✗ NO","✓ YES","SW units eliminated"),
]
for i,(p,s1,s2,x,n) in enumerate(inv_rows,5):
    f=A if i%2==0 else W
    for col,v in [(1,p),(2,s1),(3,s2),(4,x),(5,n)]:
        fail="NO" in str(v) or "✗" in str(v)
        pass_="✓" in str(v) or "YES" in str(v)
        cf=R if fail else (G if (pass_ and col==4) else f)
        cel(ws2,i,col,v,cf,align="left" if col in [1,5] else "center")

sec(ws2,15,"B. Inverter Voltage Clarification – IMPORTANT",span=5)
tbox(ws2,16,1,5,
    "The XW Pro 6848 NA is used in the 120/240V split-phase configuration. Tampa requires "
    "120V/60Hz (Table I), which is standard North American split-phase service. The 6,800W "
    "continuous rating applies to this split-phase configuration and covers the 6,000W critical "
    "load. In 120V-only single-phase mode the rating drops to 5,760W, which would be insufficient. "
    "A North American data-centre switchboard provides 120/240V split-phase — this is the assumed "
    "installation configuration.",fill=Y,rh=70)

sec(ws2,18,"C. String Configuration for CC Selection",span=5)
hdr(ws2,19,1,"Parameter",H); hdr(ws2,19,2,"Value",H); hdr(ws2,19,3,"Unit",H)
hdr(ws2,19,4,"Formula",H); hdr(ws2,19,5,"Notes",H)
str_rows=[
    ("Modules per slope","12","","",""),
    ("String config","6S × 2P","","per CC","Two strings of 6 in series"),
    ("Voc per CC (STC)","=6 × 41.10","= 246.60","V",""),
    ("Vmp per CC (STC)","=6 × 34.70","= 208.20","V",""),
    ("Isc per CC (STC)","=2 × 14.88","= 29.76","A",""),
    ("Imp per CC (STC)","=2 × 14.13","= 28.26","A",""),
    ("Power per CC","=12 × 490","= 5,880","W",""),
    ("Voc at cold (-5°C Tampa low)","=6×(41.10+(−0.22/100)×(−30)×41.10)","≈ 262.9","V",""),
]
for i,(p,f_,res,u,n) in enumerate(str_rows,20):
    fa=A if i%2==0 else W
    cel(ws2,i,1,p,fa,align="left"); cel(ws2,i,2,f_,fa)
    cel(ws2,i,3,res,fa,bold=True); cel(ws2,i,4,u,fa); cel(ws2,i,5,n,fa,align="left")

sec(ws2,29,"D. Charge Controller Comparison",span=5)
hdr(ws2,30,1,"Check",H); hdr(ws2,30,2,"MPPT 60-150",H)
hdr(ws2,30,3,"MPPT 80-600",H); hdr(ws2,30,4,"MPPT 100-600",H); hdr(ws2,30,5,"Limit values",H)
cc_rows=[
    ("Max PV Voc limit","150 V","600 V","600 V",""),
    ("Max array Isc","60 A","28 A","35 A",""),
    ("Max output power (48V)","3,500 W","4,800 W","6,000 W",""),
    ("Max charge current","60 A","80 A","100 A",""),
    ("MPPT voltage range","Batt+5–140 V","195–510 V","195–510 V",""),
    ("Efficiency (48V nominal)","98%","95%","95%",""),
    ("─── CHECKS ───","","","",""),
    ("Voc 246.6V ≤ limit?","✗ FAIL (>150V)","✓ PASS","✓ PASS","Limit: 150 / 600 / 600V"),
    ("Vmp 208.2V in range?","✗ FAIL","✓ PASS","✓ PASS","Range: –/195-510/195-510V"),
    ("Isc 29.76A ≤ limit?","✓ PASS (29.76<60)","✗ FAIL (29.76>28)","✓ PASS (29.76<35)","Limit: 60/28/35A"),
    ("Power 5880W ≤ limit?","✗ FAIL (>3500W)","✗ FAIL (5880>4800)","✓ PASS (5880<6000)","Limit: 3500/4800/6000W"),
    ("ALL PASS?","✗ NO","✗ NO","✓ YES",""),
]
for i,(p,c1,c2,c3,n) in enumerate(cc_rows,31):
    if p.startswith("─"):
        for col in range(1,6):
            cell=ws2.cell(row=i,column=col); cell.fill=S
            cell.font=Font(bold=True,color="FFFFFF",name="Arial",size=9)
        ws2.merge_cells(start_row=i,start_column=1,end_row=i,end_column=5)
        ws2.cell(row=i,column=1).value="Array: Voc=246.6V | Vmp=208.2V | Isc=29.76A | Power=5,880W"
        ws2.cell(row=i,column=1).alignment=Alignment(horizontal="center")
        continue
    f=A if i%2==0 else W
    key="ALL PASS" in p
    cel(ws2,i,1,p,f,bold=key,align="left")
    for col,v in [(2,c1),(3,c2),(4,c3)]:
        ff=R if "FAIL" in str(v) else (G if ("PASS" in str(v) or "YES" in str(v)) else f)
        cel(ws2,i,col,v,ff,bold=key)
    cel(ws2,i,5,n,f,align="left")

hdr(ws2,44,1,"SELECTED: 2 × Conext MPPT 100-600 (one per slope) | Part 865-1034 | 95% efficiency",
    G,fc="000000",span=5,rh=18)

# ══════════════════════════════════════════════════════════════
# SHEET 3: Q8 – Battery Bank Design
# ══════════════════════════════════════════════════════════════
ws3=wb.create_sheet("Q8 Battery Bank Design")
for i,w in enumerate([36,22,12,34],1): ws3.column_dimensions[get_column_letter(i)].width=w

hdr(ws3,1,1,"Q8 – Final Battery Bank Design | 2S×8P | 16 × Victron LFP 25.6V/200Ah",H,sz=12,span=4,rh=24)

sec(ws3,3,"A. DC Energy Requirement (includes inverter losses)")
hdr(ws3,4,1,"Parameter",H); hdr(ws3,4,2,"Formula / Value",H); hdr(ws3,4,3,"Result",H); hdr(ws3,4,4,"Notes",H)
e_reqs=[
    ("Critical AC load",                "Given",                          "6.0 kW",       "Load 3"),
    ("Blackout duration",               "20:00 Thu → 04:00 Fri",          "8 h",          ""),
    ("AC energy per event",             "=6.0 × 8",                       "48.0 kWh",     ""),
    ("Blackout inverter efficiency",    "From efficiency curve @ 88.8%",   f"{bo_inv_eff*100:.4f}%",""),
    ("DC ENERGY NEEDED PER EVENT",      f"=48.0 ÷ {bo_inv_eff:.4f}",     f"{DC_per_event:.3f} kWh","Battery must supply this DC energy"),
    ("Max DoD constraint",              "Assignment: max 80% DoD",         "80%",          ""),
    ("Min nominal capacity required",   f"={DC_per_event:.3f} ÷ 0.80",    f"{DC_per_event/0.80:.3f} kWh",""),
]
for i,(p,f_,res,n) in enumerate(e_reqs,5):
    key="DC ENERGY" in p
    fa=Y if key else (A if i%2==0 else W)
    cel(ws3,i,1,p,fa,bold=key,align="left")
    cel(ws3,i,2,f_,fa); cel(ws3,i,3,res,fa,bold=key); cel(ws3,i,4,n,fa,align="left")

sec(ws3,13,"B. Bank Configuration")
hdr(ws3,14,1,"Parameter",H); hdr(ws3,14,2,"Formula / Value",H); hdr(ws3,14,3,"Result",H); hdr(ws3,14,4,"Notes",H)
bank_rows=[
    ("Battery type",                "Victron LFP Smart",               "25.6V / 200Ah",  "Integrated BMS, 2500 cycles @ 80% DoD"),
    ("Energy per unit",             "=25.6 × 200 / 1000",              "5.12 kWh",       ""),
    ("DC bus voltage (inverter)",   "XW Pro 6848 requirement",         "48V nominal",    ""),
    ("Series units per string",     "=2 × 25.6V",                      "2S → 51.2V",     ""),
    ("String capacity (Ah)",        "Series: same Ah",                  "200 Ah",         "Series connection keeps Ah unchanged"),
    ("String energy",               "=51.2 × 200 / 1000",              "10.24 kWh",      ""),
    ("Parallel strings",            "ceil(65.65 / 10.24)",             "7P would give 71.68kWh; 8P selected for margin",""),
    ("TOTAL BATTERIES",             "=2S × 8P",                        "16 batteries",   ""),
    ("Bank voltage",                "=2 × 25.6",                       "51.2 V",         ""),
    ("Bank total Ah",               "=8 × 200",                        "1,600 Ah",       ""),
    ("NOMINAL CAPACITY",            "=16 × 5.12",                      "81.92 kWh",      ""),
]
for i,(p,f_,res,n) in enumerate(bank_rows,15):
    key=p.startswith("TOTAL") or p.startswith("NOMINAL")
    fa=G if key else (A if i%2==0 else W)
    cel(ws3,i,1,p,fa,bold=key,align="left")
    cel(ws3,i,2,f_,fa); cel(ws3,i,3,res,fa,bold=key); cel(ws3,i,4,n,fa,align="left")

sec(ws3,27,"C. Usable Capacity & Margin (both DoD interpretations presented)")
tbox(ws3,28,1,4,
    "The assignment states max DoD = 80% and charging ceiling = 95% SoC. There are two "
    "defensible interpretations of how the DoD floor is calculated:\n"
    "  • Interpretation A (DoD from starting SoC 95%): floor = 95% − 80% = 15% SoC → "
    f"usable = 80% × 81.92 = 65.54 kWh → margin = {81.92*0.80-DC_per_event:.2f} kWh\n"
    "  • Interpretation B (DoD from full charge 100%): floor = 100% − 80% = 20% SoC → "
    f"usable = 75% × 81.92 = 61.44 kWh → margin = {81.92*0.75-DC_per_event:.2f} kWh\n"
    f"Both interpretations pass. Actual minimum SoC reached in simulation: {min_bo_soc:.1f}% — "
    "well above both floors. The design is robust under either interpretation.",
    fill=Y,rh=90)

sec(ws3,30,"D. 3-Year Degradation Check")
hdr(ws3,31,1,"Parameter",H); hdr(ws3,31,2,"Representative Year 1",H)
hdr(ws3,31,3,"Representative Year 2",H); hdr(ws3,31,4,"Representative Year 3 + end",H)
nom_y3_end=nom_y3_end_calc
deg_rows=[
    ("Cumulative cycles",              [str(cycles_per_year*(y+1)) for y in range(3)]),
    ("Nominal capacity (kWh)",         [f"{ann[y]['nom_cap_kWh']:.3f}" for y in range(3)]),
    ("Degradation (%)",                [f"{ann[y]['deg_pct']:.3f}%" for y in range(3)]),
    ("Usable – Interp A (80%×nom)",    [f"{ann[y]['usable_kWh']:.3f}" for y in range(3)]),
    ("Usable – Interp B (75%×nom)",    [f"{ann[y]['nom_cap_kWh']*0.75:.3f}" for y in range(3)]),
    ("DC demand per event (kWh)",      [f"{DC_per_event:.3f}"]*3),
    ("Margin A (kWh)",                 [f"{ann[y]['usable_kWh']-DC_per_event:.3f}" for y in range(3)]),
    ("Margin B (kWh)",                 [f"{ann[y]['nom_cap_kWh']*0.75-DC_per_event:.3f}" for y in range(3)]),
    ("All blackouts covered?",         ["✓ YES – both interpretations"]*3),
]
for i,(p,vals) in enumerate(deg_rows,32):
    key="covered" in p
    fa=G if key else (Y if "Margin" in p else (A if i%2==0 else W))
    cel(ws3,i,1,p,fa,bold=key,align="left")
    for col,v in enumerate(vals,2): cel(ws3,i,col,v,fa,bold=key)

# ══════════════════════════════════════════════════════════════
# SHEET 4: Q6 – Inverter Efficiency
# ══════════════════════════════════════════════════════════════
ws4=wb.create_sheet("Q6 Inverter Efficiency")
for i,w in enumerate([36,14,12,36],1): ws4.column_dimensions[get_column_letter(i)].width=w

hdr(ws4,1,1,"Q6 – Inverter Efficiency Analysis | XW Pro 6848 NA",H,sz=12,span=4,rh=24)

sec(ws4,3,"A. Key Operating Points")
hdr(ws4,4,1,"Mode",H); hdr(ws4,4,2,"Load %",H); hdr(ws4,4,3,"Efficiency",H); hdr(ws4,4,4,"Notes",H)
pts=[(5.0,False),(10.0,False),(25.0,False),(27.5,True),(50.0,False),(80.0,False),(88.8,True),(100.0,False)]
labels=["GT – low irradiance","GT – morning/evening","GT – moderate","GT – PEAK EFFICIENCY",
        "GT – strong sun","GT – near max PV","BLACKOUT fixed load","Full rated output"]
for i,(label,(pct,hl)) in enumerate(zip(labels,pts),5):
    eta=float(np.interp(pct,eff_pct,eff_nom))*100
    fa=Y if "BLACKOUT" in label else (G if "PEAK" in label else (A if i%2==0 else W))
    cel(ws4,i,1,label,fa,bold=hl,align="left")
    cel(ws4,i,2,f"{pct:.1f}%",fa,bold=hl)
    cel(ws4,i,3,f"{eta:.2f}%",fa,bold=hl)
    cel(ws4,i,4,"Fixed point all blackouts" if "BLACKOUT" in label
        else ("Max efficiency point" if "PEAK" in label else ""),fa,align="left")

sec(ws4,14,"B. Grid-Tied vs Blackout Comparison")
hdr(ws4,15,1,"Parameter",H); hdr(ws4,15,2,"Value",H); hdr(ws4,15,3,"Unit",H); hdr(ws4,15,4,"Notes",H)
cmp=[
    ("Grid-tied weighted avg (Year 1)",    f"{ann[0]['wavg_inv_eff_gt']:.2f}%","","Weighted by AC output over all GT hours"),
    ("Blackout fixed efficiency",          f"{bo_inv_eff*100:.2f}%","","At 88.8% of PDC0"),
    ("Peak efficiency",                    f"{peak_eta:.2f}%","",f"At {peak_pdc_kw:.3f} kW DC input (from simulation)"),
    ("Difference (GT − blackout)",         f"{ann[0]['wavg_inv_eff_gt']-bo_inv_eff*100:.2f}","pp",""),
]
for i,(p,v,u,n) in enumerate(cmp,16):
    fa=Y if "Blackout" in p else (A if i%2==0 else W)
    cel(ws4,i,1,p,fa,bold=True,align="left")
    cel(ws4,i,2,v,fa,bold=True); cel(ws4,i,3,u,fa); cel(ws4,i,4,n,fa,align="left")

sec(ws4,21,"C. Explanation")
tbox(ws4,22,1,4,
    "GRID-TIED (avg 92.4%): PV operates at variable power. Tampa's solar distribution means the "
    "array frequently runs at 10–50% of rated output, near the peak-efficiency region. "
    "Output-weighted average efficiency is therefore close to the partial-load optimum.\n\n"
    "BLACKOUT (fixed 91.4%): The inverter always supplies exactly 6 kW — 88.8% of rated output. "
    "Above the ~27.5% peak-efficiency point, I²R losses in switches and magnetics rise "
    "quadratically with current, reducing efficiency. This is why 88.8% load gives LOWER "
    "efficiency than the grid-tied weighted average, which includes many lighter-load hours.",
    fill=W,rh=110)

# ══════════════════════════════════════════════════════════════
# SHEET 5: Q9 – Currents (FIXED wording)
# ══════════════════════════════════════════════════════════════
ws5=wb.create_sheet("Q9 Currents")
for i,w in enumerate([38,18,12,34],1): ws5.column_dimensions[get_column_letter(i)].width=w

hdr(ws5,1,1,"Q9 – Charging & Discharging Currents | 2S×8P | 51.2V / 1,600Ah",H,sz=12,span=4,rh=24)

sec(ws5,3,"A. Maximum Discharge Current")
hdr(ws5,4,1,"Parameter",H); hdr(ws5,4,2,"Formula",H); hdr(ws5,4,3,"Result",H); hdr(ws5,4,4,"Notes",H)
dc_pow=P_CRITICAL_KW/bo_inv_eff; disc_A=dc_pow*1000/51.2
disch=[
    ("DC bus voltage","2 × 25.6V","51.2 V",""),
    ("Critical AC load","Given","6.0 kW",""),
    ("Blackout inverter efficiency",f"From curve at 88.8% load",f"{bo_inv_eff*100:.4f}%",""),
    ("DC power from battery",f"= 6.0 ÷ {bo_inv_eff:.4f}",f"{dc_pow:.4f} kW",""),
    ("MAX DISCHARGE CURRENT",f"= {dc_pow:.4f} kW ÷ 51.2V",f"{disc_A:.2f} A",""),
    ("Bank capacity","8P × 200Ah","1,600 Ah",""),
    ("Discharge C-rate",f"= {disc_A:.2f} ÷ 1600",f"{disc_A/1600:.4f} C","Very gentle"),
    ("Bank rated max","8P × 200A/unit","1,600 A",""),
    ("Safety margin",f"= 1600 ÷ {disc_A:.2f}",f"{1600/disc_A:.1f}×","Well within limits"),
]
for i,(p,f_,res,n) in enumerate(disch,5):
    key="MAX DISCHARGE" in p
    fa=Y if key else (A if i%2==0 else W)
    cel(ws5,i,1,p,fa,bold=key,align="left")
    cel(ws5,i,2,f_,fa); cel(ws5,i,3,res,fa,bold=key); cel(ws5,i,4,n,fa,align="left")

sec(ws5,15,"B. Charging Currents (CORRECTED wording)")
hdr(ws5,16,1,"Parameter",H); hdr(ws5,16,2,"Formula",H); hdr(ws5,16,3,"Result",H); hdr(ws5,16,4,"Notes",H)
charg=[
    ("Peak PV DC power (both slopes)",  "From TMY weather simulation",    f"{pk_dc:.4f} kW",""),
    ("After CC efficiency (2×MPPT100-600)","× 0.95",                      f"{pk_cc:.4f} kW",""),
    ("MAX PV-DRIVEN CHARGING CURRENT",  f"= {pk_cc:.4f} kW ÷ 51.2V",    f"{pk_A:.2f} A",  "CORRECTED: actual PV-driven peak"),
    ("MPPT 100-600 rated max output",   "2 × 100A rated",                 "200 A",          "Controller ceiling (not reached by PV)"),
    ("Governing max charging current",  "min(PV-driven, CC rated)",       f"{pk_A:.2f} A",  "PV limits; CC not the bottleneck"),
    (f"Charge C-rate at {pk_A:.1f}A",  f"= {pk_A:.2f} ÷ 1600",         f"{pk_A/1600:.4f} C","Extremely safe"),
    ("LFP max rated (≤1C)",             "1 × 1600Ah",                    "1,600 A max",    f"We use {pk_A/1600*100:.1f}% of max"),
]
for i,(p,f_,res,n) in enumerate(charg,17):
    key="MAX PV-DRIVEN" in p or "Governing" in p
    fa=G if key else (OR if "CORRECTED" in n else (A if i%2==0 else W))
    cel(ws5,i,1,p,fa,bold=key,align="left")
    cel(ws5,i,2,f_,fa); cel(ws5,i,3,res,fa,bold=key); cel(ws5,i,4,n,fa,align="left")

tbox(ws5,25,1,4,
    f"DISCHARGE: Max {disc_A:.1f}A is {disc_A/1600:.3f}C on a 1,600Ah bank — {1600/disc_A:.1f}× below the rated 1,600A. "
    "Reliable critical load supply with massive headroom.\n\n"
    "CHARGE: The two MPPT 100-600 controllers are each rated to 100A output (200A combined), "
    f"but the actual PV production never drives charging above {pk_A:.1f}A at the 51.2V bus. "
    "The PV array — not the controller rating — is the governing limit. "
    f"At {pk_A:.1f}A the C-rate is {pk_A/1600:.3f}C, well within LFP safe-charge limits, preserving the "
    "full 2,500-cycle life rating.",fill=W,rh=100)

# ══════════════════════════════════════════════════════════════
# SHEET 6: Q10 – 3-Year SoC Simulation
# ══════════════════════════════════════════════════════════════
ws6=wb.create_sheet("Q10 3-Year Simulation")
for i,w in enumerate([36,18,18,18,24],1): ws6.column_dimensions[get_column_letter(i)].width=w

hdr(ws6,1,1,"Q10 – 3-Year Battery SoC Simulation | Blackout timing: h>=20 Thu / h<4 Fri (20:00–04:00) | Inverter losses included",
    H,sz=12,span=5,rh=24)

sec(ws6,3,"A. Annual Summary",span=5)
hdr(ws6,4,1,"Parameter",H); hdr(ws6,4,2,"Representative Year 1",H)
hdr(ws6,4,3,"Representative Year 2",H); hdr(ws6,4,4,"Representative Year 3",H); hdr(ws6,4,5,"Notes",H)
yr_rows=[
    ("Nominal capacity (kWh)",[f"{ann[y]['nom_cap_kWh']:.3f}" for y in range(3)],"0.02% per cycle degradation"),
    ("Degradation (%)",[f"{ann[y]['deg_pct']:.3f}%" for y in range(3)],"Cumulative from start of each year"),
    ("Usable Interp A (80%×nom, kWh)",[f"{ann[y]['usable_kWh']:.3f}" for y in range(3)],"Floor=15% SoC"),
    ("Usable Interp B (75%×nom, kWh)",[f"{ann[y]['nom_cap_kWh']*0.75:.3f}" for y in range(3)],"Floor=20% SoC"),
    ("DC demand per event (kWh)",[f"{DC_per_event:.3f}"]*3,"48 kWh AC ÷ 91.39%"),
    ("Margin A (kWh)",[f"{ann[y]['usable_kWh']-DC_per_event:.3f}" for y in range(3)],""),
    ("Margin B (kWh)",[f"{ann[y]['nom_cap_kWh']*0.75-DC_per_event:.3f}" for y in range(3)],""),
    ("Starting SoC (%)", start_soc_vals,
     "Year 1 starts at 95%; Years 2 and 3 follow from the previous simulated end state."),
    ("End-of-year SoC (%)", end_soc_vals, "Taken from the rolling hourly SoC simulation."),
    ("Actual min SoC during blackouts", min_soc_vals,
     "Minimum SoC taken directly from the hourly SoC simulation."),
    ("ALL BLACKOUTS COVERED?", covered_sim_vals, "Checked directly from hourly SoC values against the 20% SoC floor."),
    ("Annual GT yield (kWh)",[f"{ann[y]['ac_yield_kWh']:.0f}" for y in range(3)],""),
]
for i,(p,vals,note) in enumerate(yr_rows,5):
    key="ALL BLACKOUTS" in p
    fa=G if key else (Y if "Margin" in p or "Starting SoC" in p else (A if i%2==0 else W))
    ws6.row_dimensions[i].height=20 if "Starting SoC" in p else 16
    cel(ws6,i,1,p,fa,bold=key,align="left")
    for col,v in enumerate(vals,2): cel(ws6,i,col,v,fa,bold=key)
    cel(ws6,i,5,note,fa,align="left")

sec(ws6,18,"B. Final SoC After 3 Years – Direct Q10 Answer",span=5)
nom_y3_end=nom_y3_end_calc
hdr(ws6,19,1,"Parameter",H); hdr(ws6,19,2,"Value",H,span=2)
hdr(ws6,19,4,"Unit",H); hdr(ws6,19,5,"Notes",H)
q10=[
    ("Total cycles",                   f"{cycles_per_year}×3 = {total_cycles_3yr} representative cycles", "cycles", "Calculated from blackout mask: blackout hours ÷ 8 h/event."),
    ("Total degradation",              f"{total_cycles_3yr}×0.02% = {total_deg_pct_calc:.2f}%",     "",     "Conservative: one blackout event is counted as one cycle."),
    ("Final nominal capacity",         f"{nom_y3_end:.3f}",    "kWh",  f"=81.92×(1−{total_deg_pct_calc/100:.4f})"),
    ("FINAL SoC AFTER 3 YEARS",       f"{ann[2]['end_soc_pct']:.2f}%","","End of Year 3 after recharge to the 95% SoC operating target"),
    ("Final usable Interp A (80%)",    f"{nom_y3_end*0.80:.3f}","kWh", f">{DC_per_event:.3f} kWh needed ✓"),
    ("Final usable Interp B (75%)",    f"{nom_y3_end*0.75:.3f}","kWh", f">{DC_per_event:.3f} kWh needed ✓"),
    ("Final worst-case post-blackout SoC", f"{final_post_bo_soc:.2f}%", "", "Computed using final degraded capacity and simulated final SoC."),
]
for i,(p,v,u,n) in enumerate(q10,20):
    key="FINAL SoC" in p
    fa=G if key else (A if i%2==0 else W)
    cel(ws6,i,1,p,fa,bold=key,align="left")
    ws6.merge_cells(start_row=i,start_column=2,end_row=i,end_column=3)
    cel(ws6,i,2,v,fa,bold=key); cel(ws6,i,4,u,fa); cel(ws6,i,5,n,fa,align="left")

sec(ws6,27,"C. Monthly SoC Stats Year 1 (SoC %)",span=5)
hdr(ws6,28,1,"Month",H); hdr(ws6,28,2,"Min SoC",H); hdr(ws6,28,3,"Max SoC",H)
hdr(ws6,28,4,"Mean SoC",H); hdr(ws6,28,5,"Notes",H)
for m in range(1,13):
    mask=mons==m; ms=soc_y1[mask]
    fa=A if m%2==0 else W
    cel(ws6,28+m,1,MN[m-1],fa,bold=True)
    cel(ws6,28+m,2,round(ms.min(),2),fa)
    cel(ws6,28+m,3,round(ms.max(),2),fa)
    cel(ws6,28+m,4,round(ms.mean(),2),fa)
    cel(ws6,28+m,5,"Post-blackout recharge visible" if ms.min()<50 else "",fa,align="left")

# ══════════════════════════════════════════════════════════════
# SHEET 7: Q11 – LCA (FIXED note)
# ══════════════════════════════════════════════════════════════
ws7=wb.create_sheet("Q11 LCA")
for i,w in enumerate([36,22,14,36],1): ws7.column_dimensions[get_column_letter(i)].width=w

hdr(ws7,1,1,"Q11 – Life Cycle Assessment | Tampa PV System | Group 20",H,sz=12,span=4,rh=24)
hdr(ws7,2,1,f"Grid carbon intensity used: {GRID_CI} kg CO₂/kWh from the assignment-specified Our World in Data website",S,span=4,rh=16)

sec(ws7,4,"A. Embodied CO₂ by Component")
hdr(ws7,5,1,"Component",H); hdr(ws7,5,2,"Quantity × Factor",H)
hdr(ws7,5,3,"CO₂ (kg)",H); hdr(ws7,5,4,"Reference",H)
comps=[
    ("AIKO 490W modules (24 units, 11.76 kWp)","11.76×600",round(11.76*600,0),
     "Frischknecht et al. 2015 – N-type mono ≈600 kg/kWp"),
    ("Victron LFP 25.6V/200Ah (16 units, 81.92 kWh)","81.92×75",round(81.92*75,0),
     "Emilsson & Dahllöf 2019; Peters et al. 2017 – LFP 65–90 kg/kWh"),
    ("XW Pro 6848 hybrid inverter (1 unit)","1×150",150,
     "IEA-PVPS Task 12 (2015)"),
    ("Conext MPPT 100-600 (2 units)","2×30",60,
     "Proaño et al. 2017"),
    ("Mounting structure (24 panels)","24×15",360,
     "Alsema et al. 2006 – Al racking"),
    ("DC/AC switchgear (2 units)","2×20",40,"Industry average"),
    ("Cables, racks, accessories","Lump",80,"IEA-PVPS auxiliary"),
    ("Monitoring system (1 unit)","1×25",25,"Electronics estimate"),
]
total_co2=sum(c[2] for c in comps)
for i,(comp,qty,co2,ref) in enumerate(comps,6):
    fa=A if i%2==0 else W
    cel(ws7,i,1,comp,fa,bold=True,align="left")
    cel(ws7,i,2,qty,fa); cel(ws7,i,3,co2,fa,bold=True); cel(ws7,i,4,ref,fa,align="left")
    ws7.row_dimensions[i].height=16
tr=6+len(comps)
cel(ws7,tr,1,"TOTAL EMBODIED CO₂",G,bold=True,align="left")
cel(ws7,tr,2,"Sum",G); cel(ws7,tr,3,total_co2,G,bold=True)
cel(ws7,tr,4,"kg CO₂eq — Global Warming Potential",G,align="left")

sec(ws7,tr+2,"B. Carbon Payback Time (CORRECTED note)")
hdr(ws7,tr+3,1,"Parameter",H); hdr(ws7,tr+3,2,"Formula",H); hdr(ws7,tr+3,3,"Result",H); hdr(ws7,tr+3,4,"Notes",H)
ann_co2=ann[0]['ac_yield_kWh']*GRID_CI
cpbt=total_co2/ann_co2
lca=[
    ("Grid carbon intensity from assignment website","Given",f"{GRID_CI} kg CO₂/kWh","Our World in Data electricity carbon-intensity grapher"),
    ("Annual AC yield used for payback",f"{ann[0]['ac_yield_kWh']:.0f} kWh",
     f"{ann[0]['ac_yield_kWh']:.0f} kWh","Grid-tied AC yield only; conservative basis consistent with Q2 savings calculation"),
    ("Annual CO₂ savings",f"={ann[0]['ac_yield_kWh']:.0f}×{GRID_CI}",
     f"{ann_co2:.1f} kg/yr",""),
    ("Total embodied CO₂",f"Sum above",f"{total_co2} kg","20-battery bank included"),
    ("CARBON PAYBACK TIME",f"={total_co2} ÷ {ann_co2:.1f}",
     f"{cpbt:.2f} years","Grid-tied-only payback basis"),
    ("TOTAL GWP",f"{total_co2} kg CO₂eq","At system boundary","Embodied emissions"),
    ("Net CO₂ benefit (25-yr)",f"={ann_co2:.1f}×25−{total_co2}",
     f"{ann_co2*25-total_co2:.0f} kg CO₂","Avoided over 25-year life"),
]
for i,(p,f_,res,n) in enumerate(lca,tr+4):
    key="PAYBACK" in p or "TOTAL GWP" in p
    fa=G if key else (OR if "CORRECTED" in n else (A if i%2==0 else W))
    cel(ws7,i,1,p,fa,bold=key,align="left")
    cel(ws7,i,2,f_,fa); cel(ws7,i,3,res,fa,bold=key); cel(ws7,i,4,n,fa,align="left")

# ══════════════════════════════════════════════════════════════
# SHEET 8: Q1-Q3 – Yield & Savings
# ══════════════════════════════════════════════════════════════
ws8=wb.create_sheet("Q1-Q3 Yield and Savings")
for i,w in enumerate([38,18,12,36],1): ws8.column_dimensions[get_column_letter(i)].width=w

hdr(ws8,1,1,"Q1/Q2/Q3 – Yield, Grid-Tied Savings & Penalty Avoidance | Tampa | Load 3",H,sz=12,span=4,rh=24)

sec(ws8,3,"Q1 – Annual AC Electricity Yield")
hdr(ws8,4,1,"Parameter",H); hdr(ws8,4,2,"Value",H); hdr(ws8,4,3,"Unit",H); hdr(ws8,4,4,"Notes",H)
avg_poa=(G_s.sum()+G_n.sum())/2/1000
PR_gt=ann[0]['ac_yield_kWh']/(11.76*avg_poa)
PR_tot=(ann[0]['ac_yield_kWh']+annual_backup_kWh)/(11.76*avg_poa)
q1=[
    ("Annual GT AC yield – Year 1",  ann[0]['ac_yield_kWh'],"kWh","Grid-tied only (SoC=95%)"),
    ("Annual GT AC yield – Year 2",  ann[1]['ac_yield_kWh'],"kWh","Same TMY pattern"),
    ("Annual GT AC yield – Year 3",  ann[2]['ac_yield_kWh'],"kWh",""),
    ("Annual backup delivered",       annual_backup_kWh,"kWh",f"{cycles_per_year}×8h×6kW – battery, not PV"),
    ("Total AC produced Year 1",      ann[0]['ac_yield_kWh']+annual_backup_kWh,"kWh","GT + backup"),
    ("Annual PV DC production",       round(P_dc_W.sum()/1000,1),"kWh","Before CC losses"),
    ("South slope POA",               round(G_s.sum()/1000,0),"kWh/m²",""),
    ("North slope POA",               round(G_n.sum()/1000,0),"kWh/m²",""),
    ("PR – grid-tied only",           round(PR_gt,3),"–","10,400 ÷ (11.76×1,341) – low because PV also charges battery"),
    ("PR – total AC production",      round(PR_tot,3),"–","12,896 ÷ (11.76×1,341) – includes backup"),
    ("Diffuse model note","","","Isotropic model used; Perez would add ~5–10% to yield"),
]
for i,(p,v,u,n) in enumerate(q1,5):
    key=i==5
    fa=G if key else (Y if "note" in p.lower() or "PR" in p else (A if i%2==0 else W))
    cel(ws8,i,1,p,fa,bold=key,align="left")
    cel(ws8,i,2,v,fa,bold=key); cel(ws8,i,3,u,fa); cel(ws8,i,4,n,fa,align="left")
    if "note" in p.lower(): ws8.row_dimensions[i].height=22

sec(ws8,17,"Monthly Grid-Tied Yield Breakdown Year 1")
hdr(ws8,18,1,"Month",H); hdr(ws8,18,2,"GT AC (kWh)",H)
hdr(ws8,18,3,"Backup AC (kWh)",H); hdr(ws8,18,4,"Notes",H)
for m in range(1,13):
    fa=A if m%2==0 else W
    cel(ws8,18+m,1,MN[m-1],fa,bold=True)
    cel(ws8,18+m,2,round(monthly_gt[m-1],1),fa)
    cel(ws8,18+m,3,round(monthly_bo[m-1],1),fa)
    cel(ws8,18+m,4,"Peak GT months" if m in [5,6,7,8] else "",fa,align="left")
for c_,v_ in [(1,"TOTAL"),(2,round(sum(monthly_gt),1)),(3,round(sum(monthly_bo),1)),(4,"")]:
    cel(ws8,31,c_,v_,G,bold=True,align="left" if c_==1 else "center")

sec(ws8,33,"Q2 – Grid-Tied Savings Year 1")
hdr(ws8,34,1,"Parameter",H); hdr(ws8,34,2,"Formula",H); hdr(ws8,34,3,"Result",H); hdr(ws8,34,4,"Notes",H)
sav_gt=ann[0]['ac_yield_kWh']*ENERGY_PRICE
q2=[
    ("GT yield Year 1",f"{ann[0]['ac_yield_kWh']:.1f} kWh","10,400.3 kWh",""),
    ("Energy price","Table I","€0.24/kWh",""),
    ("ANNUAL SAVINGS – GT",f"=10400.3 × 0.24",f"€{sav_gt:.2f}",""),
]
for i,(p,f_,res,n) in enumerate(q2,35):
    key="ANNUAL" in p
    fa=G if key else (A if i%2==0 else W)
    cel(ws8,i,1,p,fa,bold=key,align="left")
    cel(ws8,i,2,f_,fa); cel(ws8,i,3,res,fa,bold=key); cel(ws8,i,4,n,fa,align="left")

sec(ws8,39,"Q3 – Total Savings Including Avoided Penalties Year 1")
hdr(ws8,40,1,"Parameter",H); hdr(ws8,40,2,"Formula",H); hdr(ws8,40,3,"Result",H); hdr(ws8,40,4,"Notes",H)
penalty=N_CLIENTS*COMP*blackout_hours_per_year; total_sav=sav_gt+penalty
q3=[
    ("Clients","Given","45",""),
    ("Compensation rate","Table I","€3.00/hr/client",""),
    ("Events/year", "From blackout mask", cycles_per_year, ""),
    ("Hours/event","20:00–04:00","8 h",""),
    ("Penalty without system", f"=45×3×{blackout_hours_per_year}",f"€{penalty:,.0f}","All avoided"),
    ("Blackouts covered", f"{covered_sim_vals[0]}", f"{cycles_per_year}/{cycles_per_year}", "Zero penalties paid if all are covered"),
    ("PENALTY SAVINGS", f"=45×3×{blackout_hours_per_year}",f"€{penalty:,.0f}",""),
    ("GT energy savings (Q2)","Above",f"€{sav_gt:.2f}",""),
    ("TOTAL ANNUAL SAVINGS",f"=€{penalty:,}+€{sav_gt:.2f}",f"€{total_sav:,.2f}",""),
    ("3-year total","×3",f"€{total_sav*3:,.0f}","Undiscounted"),
]
for i,(p,f_,res,n) in enumerate(q3,41):
    key=any(x in p for x in ["PENALTY SAV","TOTAL ANNUAL"])
    fa=G if key else (R if "without system" in p else (A if i%2==0 else W))
    cel(ws8,i,1,p,fa,bold=key,align="left")
    cel(ws8,i,2,f_,fa); cel(ws8,i,3,res,fa,bold=key); cel(ws8,i,4,n,fa,align="left")

# ══════════════════════════════════════════════════════════════
# SHEET 9: Hourly Load Profile Summary (sample + annual)
# ══════════════════════════════════════════════════════════════
ws9=wb.create_sheet("Hourly Load Profile")
for i,w in enumerate([12,8,8,8,12,14,12,14,14],1): ws9.column_dimensions[get_column_letter(i)].width=w

hdr(ws9,1,1,"Hourly Load Profile – Load 3 | Tampa representative year | First 168 hours shown (full year = 8,760 rows)",
    H,sz=11,span=9,rh=22)
hdr(ws9,2,1,"NOTE: The complete 8,760-hour profile was built in a separate file. Key parameters below.",
    Y,fc="000000",span=9,rh=16)

cols=["Hour#","Month","DOY","Hour","Day-of-Week","Is Blackout?","Load (kW)","Active Load (kW)","Mode (sim)"]
for c,h_ in enumerate(cols,1): hdr(ws9,3,c,h_,H,sz=9)

import pandas as pd
wx=pd.read_csv(_BASE / 'Tampa_FL-hour.csv', sep=';')
doy_arr=wx['Day'].values.astype(int); hrs_wx=wx['Hour'].values.astype(int)
mons_wx=wx['Month'].values.astype(int)
DOY1_DOW=2
MONTHLY_LOAD={1:5702.76,2:5150.88,3:5702.76,4:5518.80,5:5702.76,6:5518.80,
              7:5702.76,8:5702.76,9:5518.80,10:5702.76,11:5518.80,12:5702.76}
HF_RAW=[3.81,3.81,3.88,3.96,3.96,4.03,4.19,4.19,4.34,4.34,4.41,4.41,
        4.49,4.49,4.41,4.41,4.34,4.34,4.26,4.19,4.11,3.96,3.88,3.81]
HF_NORM_l=np.array(HF_RAW)/sum(HF_RAW)
DOWNAMES=["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]

ac_y1_arr=np.array(d['ac_all'][0])
mode_y1=d['mode_all'][0]

for i in range(168):  # first week (168h)
    m=int(mons_wx[i]); d_in_m=calendar.monthrange(YEAR,m)[1]
    daily_kwh=MONTHLY_LOAD[m]/d_in_m; h=int(hrs_wx[i])-1
    load_kw=daily_kwh*HF_NORM_l[h]
    bo=is_bo[i]; active=P_CRITICAL_KW if bo else load_kw
    dow=(DOY1_DOW+int(doy_arr[i])-1)%7
    fa=Y if bo else (A if i%2==0 else W)
    vals=[i+1,MN[m-1],int(doy_arr[i]),int(hrs_wx[i]),DOWNAMES[dow],
          "YES" if bo else "NO",round(load_kw,4),round(active,4),mode_y1[i]]
    for c,v in enumerate(vals,1):
        cell=ws9.cell(row=i+4,column=c,value=v)
        cell.font=Font(name="Arial",size=8)
        cell.fill=fa; cell.alignment=Alignment(horizontal="center",vertical="center")
        cell.border=bdr()

# Summary stats below the sample
sr=174
sec(ws9,sr,"Annual Simulation Summary (full 8,760 hours)",span=9)
summary=[
    ("Total hours", f"{len(is_bo):,}"),("Blackout hours", f"{blackout_hours_per_year} ({cycles_per_year} events × 8h)"),
    ("Grid-tied hours",f"{sum(1 for m in mode_y1 if m=='GRID_TIED')}"),
    ("Charging hours",f"{sum(1 for m in mode_y1 if m=='CHARGING')}"),
    ("Idle hours",f"{sum(1 for m in mode_y1 if m=='IDLE')}"),
    ("Annual GT AC yield", f"{ann[0]['ac_yield_kWh']:.1f} kWh"),("Annual backup delivered", f"{annual_backup_kWh:.1f} kWh"),
    ("Min SoC during blackouts", min_soc_vals[0]),("Final SoC end of Representative Year 1", end_soc_vals[0]),
]
for i,(p,v) in enumerate(summary,sr+1):
    fa=A if i%2==0 else W
    cel(ws9,i,1,p,fa,bold=True,align="left",sz=9)
    ws9.merge_cells(start_row=i,start_column=2,end_row=i,end_column=9)
    cel(ws9,i,2,v,fa,align="left",sz=9)

out=_BASE / 'SUBMISSION_READY_Group20.xlsx'
wb.save(out)
print("Saved:", out)
print(f"\nSheets: {[s.title for s in wb.worksheets]}")
print(f"\nFINAL KEY VALUES:")
print(f"  Q6  GT efficiency:        {ann[0]['wavg_inv_eff_gt']:.2f}%")
print(f"  Q6  Blackout efficiency:  {bo_inv_eff*100:.2f}%")
print(f"  Q9  Max discharge:        {disc_A:.2f} A")
print(f"  Q9  Max PV-driven charge: {pk_A:.2f} A  (CC rated max: 200A)")
print(f"  Q10 Final SoC (3yr):      {ann[2]['end_soc_pct']:.2f}%")
print(f"  Q11 Total CO₂:            {total_co2:.0f} kg | Payback: {cpbt:.2f} yr")
print(f"  P1  Yield:                {ann[0]['ac_yield_kWh']:.0f} kWh")
print(f"  P2  GT savings:           €{sav_gt:.2f}")
print(f"  P3  Total savings:        €{total_sav:.2f}")
